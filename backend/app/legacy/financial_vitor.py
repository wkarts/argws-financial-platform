from __future__ import annotations

import csv
import hashlib
import re
import shutil
import tempfile
import unicodedata
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.tenant import Company, Contract, Customer, CustomerContact, Receivable, ServiceCatalog

MAX_ARCHIVE_FILES = 2_000
MAX_UNCOMPRESSED_BYTES = 500 * 1024 * 1024


def normalized_name(value: str) -> str:
    text = unicodedata.normalize("NFKD", value or "").encode("ascii", "ignore").decode("ascii")
    return " ".join(re.sub(r"[^A-Z0-9 ]", " ", text.upper()).split())


def parse_decimal(value: Any) -> Decimal:
    if isinstance(value, Decimal):
        return value.quantize(Decimal("0.01"))
    if isinstance(value, (int, float)):
        return Decimal(str(value)).quantize(Decimal("0.01"))
    raw = str(value or "").replace("R$", "").replace(" ", "")
    if "," in raw:
        raw = raw.replace(".", "").replace(",", ".")
    try:
        return Decimal(raw).quantize(Decimal("0.01"))
    except InvalidOperation:
        return Decimal("0.00")


def parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value or "").strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            pass
    return None


def competence_from_names(paths: list[Path]) -> tuple[int, int]:
    for path in paths:
        for month, year in re.findall(r"(?<!\d)(0[1-9]|1[0-2])[-_ ]?(20\d{2})(?!\d)", path.name):
            return int(year), int(month)
    dates: list[date] = []
    for path in paths:
        if path.suffix.lower() not in {".txt", ".csv"}:
            continue
        try:
            for line in path.read_text(encoding="utf-8", errors="replace").splitlines()[:200]:
                for raw in re.findall(r"\b\d{2}/\d{2}/\d{4}\b", line):
                    parsed = parse_date(raw)
                    if parsed:
                        dates.append(parsed)
        except OSError:
            continue
    if dates:
        dates.sort()
        middle = dates[len(dates) // 2]
        return middle.year, middle.month
    today = date.today()
    return today.year, today.month


@dataclass(slots=True)
class LegacyRecord:
    name: str
    amount: Decimal
    due_date: date | None = None
    billing_method: str = "MANUAL"
    tax_id: str | None = None
    email: str | None = None
    whatsapp: str | None = None
    source_files: set[str] = field(default_factory=set)


@dataclass(slots=True)
class LegacyPreview:
    competence: str
    archive_sha256: str
    records: list[LegacyRecord]
    warnings: list[str]
    source_counts: dict[str, int]

    def to_dict(self) -> dict[str, Any]:
        return {
            "competence": self.competence,
            "archive_sha256": self.archive_sha256,
            "records": [
                {
                    "name": item.name,
                    "amount": str(item.amount),
                    "due_date": item.due_date.isoformat() if item.due_date else None,
                    "billing_method": item.billing_method,
                    "tax_id": item.tax_id,
                    "email": item.email,
                    "whatsapp": item.whatsapp,
                    "source_files": sorted(item.source_files),
                }
                for item in self.records
            ],
            "warnings": self.warnings,
            "source_counts": self.source_counts,
        }


class FinancialVitorImporter:
    def _extract(self, archive: Path) -> Path:
        target = Path(tempfile.mkdtemp(prefix="financial-vitor-"))
        try:
            with zipfile.ZipFile(archive) as zf:
                members = zf.infolist()
                if len(members) > MAX_ARCHIVE_FILES:
                    raise ValueError(f"ZIP excede o limite de {MAX_ARCHIVE_FILES} arquivos.")
                total_size = sum(member.file_size for member in members)
                if total_size > MAX_UNCOMPRESSED_BYTES:
                    raise ValueError("ZIP excede o limite descompactado de 500 MB.")
                for member in members:
                    resolved = (target / member.filename).resolve()
                    if target.resolve() not in resolved.parents and resolved != target.resolve():
                        raise ValueError("ZIP contém caminho inseguro.")
                    if member.is_dir():
                        continue
                    ratio = member.file_size / max(member.compress_size, 1)
                    if member.file_size > 50 * 1024 * 1024 and ratio > 200:
                        raise ValueError("ZIP contém arquivo com taxa de compactação suspeita.")
                zf.extractall(target)
            return target
        except Exception:
            shutil.rmtree(target, ignore_errors=True)
            raise

    @staticmethod
    def _find(root: Path, pattern: str) -> Path | None:
        return next((path for path in root.rglob(pattern) if path.is_file()), None)

    @staticmethod
    def _sha256(path: Path) -> str:
        digest = hashlib.sha256()
        with path.open("rb") as stream:
            for chunk in iter(lambda: stream.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()

    def preview(self, archive: Path) -> LegacyPreview:
        root = self._extract(archive)
        try:
            paths = [path for path in root.rglob("*") if path.is_file()]
            year, month = competence_from_names(paths)
            competence = f"{year:04d}-{month:02d}"
            default_due = date(year, month, 25)
            records: dict[str, LegacyRecord] = {}
            warnings: list[str] = []
            counts: dict[str, int] = {}

            honorario = self._find(root, "HONORARIO*.xlsx")
            if honorario:
                workbook = load_workbook(honorario, data_only=True, read_only=True)
                worksheet = workbook.active
                for row in worksheet.iter_rows(min_row=3, values_only=True):
                    name = str(row[0] or "").strip()
                    amount = parse_decimal(row[1])
                    if not name or amount <= 0:
                        continue
                    records[normalized_name(name)] = LegacyRecord(
                        name=name,
                        amount=amount,
                        source_files={honorario.name},
                    )
                workbook.close()
                counts[honorario.name] = len(records)
            else:
                warnings.append("Planilha HONORARIO não localizada.")

            for path in sorted(root.rglob("boleto2*.txt")):
                count = 0
                for line in path.read_text(encoding="utf-8", errors="replace").splitlines():
                    columns = [part.strip() for part in line.split("\t")]
                    if len(columns) < 3:
                        continue
                    key = normalized_name(columns[0])
                    item = records.get(key)
                    if item is None:
                        item = LegacyRecord(name=columns[0], amount=parse_decimal(columns[2]))
                        records[key] = item
                    item.due_date = parse_date(columns[1])
                    item.billing_method = "BOLETO"
                    item.source_files.add(path.name)
                    count += 1
                counts[path.name] = count

            receipt_csv = self._find(root, "recibos*.csv")
            if receipt_csv:
                count = 0
                with receipt_csv.open(encoding="utf-8-sig", errors="replace", newline="") as stream:
                    for row in csv.reader(stream, delimiter=";"):
                        if len(row) < 3 or not parse_date(row[1]):
                            continue
                        key = normalized_name(row[0])
                        item = records.get(key)
                        if item is None:
                            item = LegacyRecord(name=row[0], amount=parse_decimal(row[2]))
                            records[key] = item
                        item.due_date = parse_date(row[1])
                        item.billing_method = "RECEIPT"
                        item.source_files.add(receipt_csv.name)
                        count += 1
                counts[receipt_csv.name] = count

            notes = self._find(root, "NOTAS*.xlsx")
            if notes:
                workbook = load_workbook(notes, data_only=True, read_only=True)
                note_count = 0
                contact_count = 0
                for worksheet in workbook.worksheets:
                    title = normalized_name(worksheet.title)
                    if "EMAIL" in title or "CONTATO" in title:
                        for row in worksheet.iter_rows(min_row=3, values_only=True):
                            name = str(row[0] or "").strip()
                            if not name:
                                continue
                            key = normalized_name(name)
                            matches = [record for candidate, record in records.items() if candidate.startswith(key) or key.startswith(candidate)]
                            if len(matches) != 1:
                                continue
                            email = str(row[1] or "").strip()
                            phone = str(row[2] or "").strip()
                            if "@" in email:
                                matches[0].email = email.lower()
                            if phone:
                                matches[0].whatsapp = "".join(ch for ch in phone if ch.isdigit()) or None
                            matches[0].source_files.add(notes.name)
                            contact_count += 1
                        continue
                    for row in worksheet.iter_rows(min_row=3, values_only=True):
                        name = str(row[0] or "").strip()
                        tax_id = "".join(ch for ch in str(row[1] or "") if ch.isdigit())
                        if not name or len(tax_id) not in {11, 14}:
                            continue
                        key = normalized_name(name)
                        matches = [record for candidate, record in records.items() if candidate.startswith(key) or key.startswith(candidate)]
                        if len(matches) == 1:
                            matches[0].tax_id = tax_id
                            matches[0].source_files.add(notes.name)
                            note_count += 1
                workbook.close()
                counts[f"{notes.name}:notas"] = note_count
                counts[f"{notes.name}:contatos"] = contact_count

            for item in records.values():
                if item.due_date is None:
                    item.due_date = default_due
                    warnings.append(f"Vencimento padrão aplicado a {item.name}.")
                if not item.tax_id:
                    warnings.append(f"CPF/CNPJ não identificado automaticamente: {item.name}.")
                if item.amount <= 0:
                    warnings.append(f"Valor inválido ou zerado: {item.name}.")
            return LegacyPreview(
                competence=competence,
                archive_sha256=self._sha256(archive),
                records=sorted(records.values(), key=lambda item: normalized_name(item.name)),
                warnings=warnings,
                source_counts=counts,
            )
        finally:
            shutil.rmtree(root, ignore_errors=True)

    async def import_into(
        self,
        session: AsyncSession,
        archive: Path,
        *,
        company_id: str,
        service_id: str,
        create_contracts: bool = True,
        create_receivables: bool = True,
    ) -> dict[str, Any]:
        preview = self.preview(archive)
        company = await session.get(Company, company_id)
        service = await session.get(ServiceCatalog, service_id)
        if company is None or service is None:
            raise ValueError("Empresa ou serviço não encontrado no tenant.")
        year, month = (int(value) for value in preview.competence.split("-", 1))
        competence_start = date(year, month, 1)
        stats: dict[str, Any] = {
            "archive_sha256": preview.archive_sha256,
            "competence": preview.competence,
            "customers_created": 0,
            "contacts_created": 0,
            "contracts_created": 0,
            "receivables_created": 0,
            "records_skipped": 0,
        }
        for item in preview.records:
            if item.amount <= 0:
                stats["records_skipped"] += 1
                continue
            customer = None
            if item.tax_id:
                customer = await session.scalar(select(Customer).where(Customer.tax_id == item.tax_id))
            if customer is None:
                customer = await session.scalar(select(Customer).where(Customer.name == item.name))
            if customer is None:
                customer = Customer(
                    person_type="PJ" if item.tax_id and len(item.tax_id) == 14 else "PF",
                    name=item.name,
                    tax_id=item.tax_id,
                    email=item.email,
                    whatsapp=item.whatsapp,
                    tags=["IMPORTADO_FINANCEIRO_VITOR"],
                    notes="Importado do pacote financeiro legado.",
                )
                session.add(customer)
                await session.flush()
                stats["customers_created"] += 1
            if (item.email or item.whatsapp) and not await session.scalar(
                select(CustomerContact.id).where(CustomerContact.customer_id == customer.id)
            ):
                session.add(
                    CustomerContact(
                        customer_id=customer.id,
                        name="Financeiro",
                        email=item.email,
                        whatsapp=item.whatsapp,
                        role="FINANCEIRO",
                        is_primary=True,
                    )
                )
                stats["contacts_created"] += 1

            digest = hashlib.sha256(normalized_name(item.name).encode()).hexdigest()[:12].upper()
            contract_code = f"LEGACY-{digest}"
            contract = await session.scalar(select(Contract).where(Contract.code == contract_code))
            due = item.due_date or date(year, month, 25)
            if create_contracts and contract is None:
                next_due = due + relativedelta(months=1)
                contract = Contract(
                    company_id=company.id,
                    customer_id=customer.id,
                    service_id=service.id,
                    code=contract_code,
                    description=f"Honorários mensais - {item.name}",
                    amount=item.amount,
                    frequency="MONTHLY",
                    interval_count=1,
                    billing_method="BOLETO_PIX" if item.billing_method == "BOLETO" else "RECEIPT",
                    due_day=due.day,
                    start_date=due,
                    next_generation_date=max(next_due - relativedelta(days=10), competence_start),
                    issue_days_before_due=10,
                    status="ACTIVE",
                    settings={"legacy_source": sorted(item.source_files)},
                )
                session.add(contract)
                await session.flush()
                stats["contracts_created"] += 1
            if create_receivables and contract is not None:
                existing = await session.scalar(
                    select(Receivable.id).where(
                        Receivable.contract_id == contract.id,
                        Receivable.competence == preview.competence,
                    )
                )
                if existing is None:
                    session.add(
                        Receivable(
                            company_id=company.id,
                            customer_id=customer.id,
                            contract_id=contract.id,
                            document_number=f"LEG-{preview.competence.replace('-', '')}-{digest}",
                            competence=preview.competence,
                            description=f"Honorários {preview.competence}",
                            issue_date=competence_start,
                            due_date=due,
                            original_amount=item.amount,
                            discount_amount=Decimal("0"),
                            interest_amount=Decimal("0"),
                            fine_amount=Decimal("0"),
                            abatement_amount=Decimal("0"),
                            paid_amount=Decimal("0"),
                            balance=item.amount,
                            status="OPEN",
                            source="LEGACY_IMPORT",
                            metadata_json={"source_files": sorted(item.source_files), "billing_method": item.billing_method},
                        )
                    )
                    stats["receivables_created"] += 1
        await session.commit()
        return stats
