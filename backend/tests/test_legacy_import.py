from __future__ import annotations

import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook

from app.legacy.financial_vitor import FinancialVitorImporter, parse_decimal


def create_fixture(root: Path) -> Path:
    source = root / "source"
    source.mkdir()

    honorario = Workbook()
    ws = honorario.active
    ws.append(["RELATÓRIO DE HONORÁRIOS"])
    ws.append(["CLIENTE", "VALOR"])
    ws.append(["EMPRESA TESTE LTDA", 1621.00])
    ws.append(["CLIENTE RECIBO", "850,50"])
    honorario.save(source / "HONORARIO 082026.xlsx")

    (source / "boleto2 082026.txt").write_text(
        "EMPRESA TESTE LTDA\t25/08/2026\t1.621,00\n",
        encoding="utf-8",
    )
    (source / "recibos 082026.csv").write_text(
        "CLIENTE RECIBO;25/08/2026;850,50\n",
        encoding="utf-8",
    )

    notas = Workbook()
    ws_notas = notas.active
    ws_notas.title = "NOTAS"
    ws_notas.append(["NOTAS"])
    ws_notas.append(["CLIENTE", "CNPJ"])
    ws_notas.append(["EMPRESA TESTE LTDA", "12.345.678/0001-90"])
    ws_contatos = notas.create_sheet("EMAIL CONTATOS")
    ws_contatos.append(["CONTATOS"])
    ws_contatos.append(["CLIENTE", "EMAIL", "WHATSAPP"])
    ws_contatos.append(["EMPRESA TESTE", "financeiro@example.com", "(75) 99999-9999"])
    notas.save(source / "NOTAS.xlsx")

    archive = root / "FINANCEIRO TESTE.zip"
    with zipfile.ZipFile(archive, "w", zipfile.ZIP_DEFLATED) as zf:
        for path in source.iterdir():
            zf.write(path, path.name)
    return archive


def test_legacy_preview_consolidates_sources(tmp_path: Path) -> None:
    preview = FinancialVitorImporter().preview(create_fixture(tmp_path))
    assert preview.competence == "2026-08"
    assert len(preview.archive_sha256) == 64
    assert len(preview.records) == 2
    first = next(item for item in preview.records if item.name == "EMPRESA TESTE LTDA")
    assert first.amount == parse_decimal("1.621,00")
    assert first.billing_method == "BOLETO"
    assert first.tax_id == "12345678000190"
    assert first.email == "financeiro@example.com"
    assert first.whatsapp == "75999999999"


def test_legacy_import_rejects_zip_path_traversal(tmp_path: Path) -> None:
    archive = tmp_path / "unsafe.zip"
    with zipfile.ZipFile(archive, "w") as zf:
        zf.writestr("../fora.txt", "não permitido")
    with pytest.raises(ValueError, match="caminho inseguro"):
        FinancialVitorImporter().preview(archive)
